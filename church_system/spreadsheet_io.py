"""Read/write helpers for Excel (.xlsx) imports."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable

from django.core.exceptions import ValidationError
from openpyxl import Workbook, load_workbook

from church_system.uploads import validate_upload

MAX_IMPORT_ROWS = 1000


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    text = re.sub(r"[^a-z0-9_]", "", text)
    return text


def read_xlsx_rows(
    uploaded,
    *,
    max_rows: int = MAX_IMPORT_ROWS,
) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Parse the first worksheet into a list of row dicts keyed by normalized headers.

    Raises ValidationError on empty/invalid files.
    """
    validate_upload(uploaded, kind="document")
    name = (getattr(uploaded, "name", "") or "").lower()
    if not name.endswith(".xlsx"):
        raise ValidationError("Member and transaction imports require an Excel .xlsx file.")

    uploaded.seek(0)
    try:
        workbook = load_workbook(uploaded, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("Could not read the Excel file. Save it as .xlsx and try again.") from exc

    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        header_cells = next(row_iter)
    except StopIteration:
        raise ValidationError("The spreadsheet is empty.")

    headers: list[str] = []
    for index, cell in enumerate(header_cells):
        key = _normalize_header(cell)
        if not key:
            key = f"column_{index + 1}"
        headers.append(key)

    if not any(h for h in headers if not h.startswith("column_")):
        raise ValidationError("Add a header row with column names in the first row.")

    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(row_iter, start=2):
        if len(rows) >= max_rows:
            raise ValidationError(f"Imports are limited to {max_rows} data rows per file.")
        if _row_is_blank(cells):
            continue
        row_dict = {}
        for index, header in enumerate(headers):
            row_dict[header] = cells[index] if index < len(cells) else None
        row_dict["_row_number"] = row_number
        rows.append(row_dict)
    workbook.close()
    if not rows:
        raise ValidationError("No data rows found below the header row.")
    return headers, rows


def _row_is_blank(cells: Iterable[Any]) -> bool:
    for cell in cells:
        if cell is None:
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def parse_excel_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValidationError(f"Unrecognized date: {text}")


def parse_decimal_amount(value: Any, *, field_label: str = "Amount") -> Decimal:
    from church_system.money import quantize_money

    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        amount = Decimal(str(value))
    else:
        text = str(value).strip().replace(",", "")
        if not text:
            return Decimal("0.00")
        try:
            amount = Decimal(text)
        except InvalidOperation as exc:
            raise ValidationError(f"{field_label} is not a valid number: {value}") from exc
    if amount < 0:
        raise ValidationError(f"{field_label} cannot be negative.")
    try:
        return quantize_money(amount)
    except ValueError as exc:
        raise ValidationError(f"{field_label} is not a valid number: {value}") from exc


def build_template_xlsx(headers: list[str], example_rows: list[list[Any]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    ws.append(headers)
    for row in example_rows or []:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
